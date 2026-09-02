from __future__ import annotations

import argparse
import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

URL = "https://live.titan007.com/index2in1.aspx?id=3"
BANGKOK = ZoneInfo("Asia/Bangkok")
HEADERS = ["赛事名称", "单场编号", "主队", "客队", "亚盘主队水位", "欧赔主胜", "亚洲盘口", "欧赔平局", "亚盘客队水位", "欧赔客胜"]


def clean_lines(text: str) -> list[str]:
    return [re.sub(r"\s+", " ", x).strip() for x in text.splitlines() if x.strip()]


def number_or_text(value: str) -> Any:
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def split_pair(text: str) -> tuple[str, str]:
    parts = clean_lines(text)
    return (parts[0] if parts else "", parts[1] if len(parts) > 1 else "")


def scrape(diagnostics_dir: Path) -> list[list[Any]]:
    diagnostics_dir.mkdir(parents=True, exist_ok=True)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--disable-dev-shm-usage"])
        page = browser.new_page(viewport={"width": 1440, "height": 1000}, locale="zh-CN", timezone_id="Asia/Bangkok")
        try:
            page.goto(URL, wait_until="domcontentloaded", timeout=90000)
            page.locator("#table_live").wait_for(state="attached", timeout=90000)
            page.wait_for_function("document.querySelectorAll('#table_live tr[id^=tr1_]').length > 0", timeout=90000)

            page.locator("#checkLet").set_checked(True)
            page.locator("#checkTotal").set_checked(False)
            page.locator("#checkEu").set_checked(True)
            page.wait_for_timeout(1200)

            page.locator("#button3").click()
            page.locator("#rbs3").click()
            try:
                page.locator('#Layer2 input[value="确定"]').click(timeout=15000)
            except PlaywrightTimeoutError:
                logging.warning("确认按钮超时，但网页通常仍已完成筛选；继续检查结果。")
            page.wait_for_timeout(1500)

            raw_rows = page.locator("#table_live tr[id^=tr1_]").evaluate_all(
                """rows => rows.filter(tr => getComputedStyle(tr).display !== 'none')
                .map(tr => [...tr.cells].map(td => (td.innerText || '').trim()))"""
            )
        except Exception:
            page.screenshot(path=str(diagnostics_dir / "failure_page.png"), full_page=True)
            (diagnostics_dir / "failure_page.html").write_text(page.content(), encoding="utf-8")
            raise
        finally:
            browser.close()

    output: list[list[Any]] = []
    for cells in raw_rows:
        if len(cells) < 12:
            continue
        event_name, single_number = split_pair(cells[1])
        home, away = " ".join(clean_lines(cells[4])), " ".join(clean_lines(cells[6]))
        if "单场" not in single_number or not home or not away:
            continue
        asian_home, euro_home = split_pair(cells[9])
        asian_handicap, euro_draw = split_pair(cells[10])
        asian_away, euro_away = split_pair(cells[11])
        output.append([event_name, single_number, home, away, number_or_text(asian_home), number_or_text(euro_home), asian_handicap, number_or_text(euro_draw), number_or_text(asian_away), number_or_text(euro_away)])
    return output


def save_excel(rows: list[list[Any]], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    now = datetime.now(BANGKOK)
    path = output_dir / f"titan007_single_{now:%Y-%m-%d_%H%M}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "单场实时数据"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row in ws.iter_rows(min_row=2, max_row=max(2, ws.max_row), min_col=1, max_col=10):
        for cell in row:
            cell.alignment, cell.border = Alignment(vertical="center"), Border(bottom=thin)
        for cell in row[1:]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for cell in (row[4], row[5], row[7], row[8], row[9]):
            cell.number_format = "0.00"

    for idx, width in enumerate([15, 12, 22, 22, 16, 14, 16, 14, 16, 14], 1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"
    table = Table(displayName="Titan007SingleData", ref=f"A1:J{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    meta = wb.create_sheet("运行信息")
    meta.append(["项目", "内容"])
    meta.append(["抓取时间（泰国）", now.strftime("%Y-%m-%d %H:%M:%S %z")])
    meta.append(["数据源", URL])
    meta.append(["筛选", "即时比分 / Crow* / 亚+欧 / 单场；十列输出"])
    meta.append(["比赛数量", len(rows)])
    meta.column_dimensions["A"].width, meta.column_dimensions["B"].width = 24, 70
    for cell in meta[1]:
        cell.fill, cell.font = header_fill, header_font
    wb.save(path)
    return path


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", default="output")
    parser.add_argument("--log-dir", default="logs")
    args = parser.parse_args()
    log_dir = Path(args.log_dir)
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / "scraper.log"
    logging.basicConfig(filename=log_path, level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", encoding="utf-8", force=True)
    try:
        rows = scrape(log_dir)
        if not rows:
            raise RuntimeError("筛选后没有抓取到单场比赛；可能当天无单场赛事，或网页结构已改变。")
        output = save_excel(rows, Path(args.output_dir))
        logging.info("成功抓取 %d 场：%s", len(rows), output)
        print(output)
        return 0
    except Exception as exc:
        logging.exception("抓取失败")
        print(f"抓取失败：{exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
